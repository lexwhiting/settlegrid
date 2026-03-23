#!/usr/bin/env python3
"""Generate SettleGrid SMOKE_TEST_TRACKER.xlsx with Dashboard + Tests sheets."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "SMOKE_TEST_TRACKER.xlsx")

# Header styling
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Type row backgrounds
TYPE_FILLS = {
    "AUTO": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "SEMI": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "MANUAL": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
}

# Status cell fills
STATUS_FILLS = {
    "PASS": PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid"),
    "FAIL": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
    "SKIP": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "PENDING": PatternFill(fill_type=None),
}

# All 72 SettleGrid tests (51 original + 21 auth)
TESTS = [
    # Phase 1: Authentication & Access Control
    ("S01", "1.1", "Authentication", "Gate Access", "MANUAL", "/"),
    ("S02", "1.2", "Authentication", "Developer Registration", "MANUAL", "/register?type=developer"),
    ("S03", "1.3", "Authentication", "Consumer Registration", "MANUAL", "/register?type=consumer"),
    ("S04", "1.4", "Authentication", "Login (existing account)", "MANUAL", "/login"),
    ("S05", "1.5", "Authentication", "Session Persistence", "MANUAL", "/dashboard"),
    ("S06", "1.6", "Authentication", "Logout", "MANUAL", "N/A (UI)"),
    # Phase 2: Developer Dashboard
    ("S07", "2.1", "Dashboard", "Overview", "AUTO", "/dashboard"),
    ("S08", "2.2", "Dashboard", "Navigation", "MANUAL", "N/A (UI)"),
    # Phase 3: Tool Management
    ("S09", "3.1", "Tool Management", "Create Tool", "MANUAL", "/tools/create"),
    ("S10", "3.2", "Tool Management", "Tool Details", "MANUAL", "/tools/[id]"),
    ("S11", "3.3", "Tool Management", "Tool Configuration", "MANUAL", "/tools/[id]/edit"),
    ("S12", "3.4", "Tool Management", "Tool Changelog", "MANUAL", "/tools/[id]/changelog"),
    # Phase 4: SDK Integration
    ("S13", "4.1", "SDK Integration", "Install SDK", "MANUAL", "npm @settlegrid/mcp"),
    ("S14", "4.2", "SDK Integration", "Initialize SDK", "AUTO", "SettleGrid.init()"),
    ("S15", "4.3", "SDK Integration", "Wrap Tool Function", "AUTO", "sg.wrap()"),
    ("S16", "4.4", "SDK Integration", "Validate Key", "AUTO", "/api/sdk/validate-key"),
    ("S17", "4.5", "SDK Integration", "Meter Usage", "AUTO", "/api/sdk/meter"),
    # Phase 5: Consumer Experience
    ("S18", "5.1", "Consumer Experience", "Tool Showcase", "MANUAL", "/tools"),
    ("S19", "5.2", "Consumer Experience", "Tool Storefront", "MANUAL", "/tools/[id]/storefront"),
    ("S20", "5.3", "Consumer Experience", "Purchase Credits", "MANUAL", "N/A (Stripe)"),
    ("S21", "5.4", "Consumer Experience", "Consumer API Keys", "MANUAL", "N/A (UI)"),
    # Phase 6: Billing & Payouts
    ("S22", "6.1", "Billing", "Stripe Connect (Developer)", "MANUAL", "/settings/connect"),
    ("S23", "6.2", "Billing", "Revenue Dashboard", "AUTO", "/analytics"),
    ("S24", "6.3", "Billing", "Payouts", "MANUAL", "/payouts"),
    ("S25", "6.4", "Billing", "Consumer Billing", "MANUAL", "N/A (UI)"),
    # Phase 7: Webhooks & Events
    ("S26", "7.1", "Webhooks", "Create Webhook Endpoint", "MANUAL", "/webhooks/add"),
    ("S27", "7.2", "Webhooks", "SSRF Protection", "SEMI", "N/A (validation)"),
    ("S28", "7.3", "Webhooks", "Webhook Delivery", "AUTO", "/api/webhooks/deliver"),
    ("S29", "7.4", "Webhooks", "Test Webhook", "MANUAL", "N/A (UI)"),
    # Phase 8: Health Checks
    ("S30", "8.1", "Health Checks", "Tool Health Monitoring", "AUTO", "/api/health"),
    ("S31", "8.2", "Health Checks", "Health Dashboard", "AUTO", "/analytics"),
    # Phase 9: Developer Profiles & Reviews
    ("S32", "9.1", "Profiles", "Public Profile", "MANUAL", "/developers/[id]"),
    ("S33", "9.2", "Profiles", "Tool Reviews", "MANUAL", "N/A (UI)"),
    # Phase 10: Alerts & Notifications
    ("S34", "10.1", "Alerts", "Consumer Alerts", "AUTO", "N/A (email)"),
    ("S35", "10.2", "Alerts", "Budget Controls", "SEMI", "N/A (SDK)"),
    # Phase 11: Audit Log
    ("S36", "11.1", "Audit", "View Audit Trail", "MANUAL", "/audit-log"),
    ("S37", "11.2", "Audit", "Export", "MANUAL", "/audit-log/export"),
    # Phase 12: Documentation
    ("S38", "12.1", "Documentation", "Docs Page", "AUTO", "/docs"),
    # Phase 13: Settings & Configuration
    ("S39", "13.1", "Settings", "Developer Settings", "MANUAL", "/settings"),
    ("S40", "13.2", "Settings", "Sandbox Mode", "SEMI", "N/A (SDK)"),
    # Phase 14: Error Handling & Edge Cases
    ("S41", "14.1", "Error Handling", "Invalid Input", "SEMI", "N/A (UI)"),
    ("S42", "14.2", "Error Handling", "Rate Limiting", "AUTO", "/api/health"),
    ("S43", "14.3", "Error Handling", "Unauthorized Access", "SEMI", "/dashboard"),
    ("S44", "14.4", "Error Handling", "Error Pages", "MANUAL", "/nonexistent"),
    # Phase 15: Performance & UX
    ("S45", "15.1", "Performance", "Page Load", "AUTO", "/dashboard"),
    ("S46", "15.2", "Performance", "Responsive Design", "MANUAL", "N/A (UI)"),
    ("S47", "15.3", "Performance", "Accessibility", "MANUAL", "N/A (UI)"),
    ("S48", "15.4", "Performance", "Brand Consistency", "AUTO", "/"),
    # Phase 16: API Direct Testing
    ("S49", "16.1", "API", "Auth", "MANUAL", "N/A (Clerk)"),
    ("S50", "16.2", "API", "Tool CRUD", "MANUAL", "N/A (DevTools)"),
    ("S51", "16.3", "API", "Health Check", "AUTO", "/api/health"),
    # ── Auth Phase A: Gate & Session (SA-prefix) ──
    ("SA01", "A.1", "Auth Gate", "Gate cookie acquired", "AUTO", "/api/gate"),
    ("SA02", "A.2", "Auth Session", "GET /api/auth/developer/me returns 200", "AUTO", "/api/auth/developer/me"),
    # ── Auth Phase B: Dashboard Pages ──
    ("SA10", "B.1", "Auth Dashboard", "Dashboard loads with auth", "AUTO", "/dashboard"),
    ("SA11", "B.2", "Auth Dashboard", "Tools page loads with auth", "AUTO", "/dashboard/tools"),
    ("SA12", "B.3", "Auth Dashboard", "Analytics page loads with auth", "AUTO", "/dashboard/analytics"),
    ("SA13", "B.4", "Auth Dashboard", "Health page loads with auth", "AUTO", "/dashboard/health"),
    ("SA14", "B.5", "Auth Dashboard", "Payouts page loads with auth", "AUTO", "/dashboard/payouts"),
    ("SA15", "B.6", "Auth Dashboard", "Referrals page loads with auth", "AUTO", "/dashboard/referrals"),
    ("SA16", "B.7", "Auth Dashboard", "Fraud page loads with auth", "AUTO", "/dashboard/fraud"),
    ("SA17", "B.8", "Auth Dashboard", "Reputation page loads with auth", "AUTO", "/dashboard/reputation"),
    ("SA18", "B.9", "Auth Dashboard", "Webhooks page loads with auth", "AUTO", "/dashboard/webhooks"),
    ("SA19", "B.10", "Auth Dashboard", "Audit log page loads with auth", "AUTO", "/dashboard/audit-log"),
    ("SA20", "B.11", "Auth Dashboard", "Settings page loads with auth", "AUTO", "/dashboard/settings"),
    ("SA21", "B.12", "Auth Dashboard", "Consumer page loads with auth", "AUTO", "/consumer"),
    # ── Auth Phase C: API Routes ──
    ("SA30", "C.1", "Auth API", "API: tools", "AUTO", "/api/tools"),
    ("SA31", "C.2", "Auth API", "API: developer stats", "AUTO", "/api/dashboard/developer/stats"),
    ("SA32", "C.3", "Auth API", "API: audit log", "AUTO", "/api/audit-log"),
    ("SA33", "C.4", "Auth API", "API: payouts", "AUTO", "/api/payouts"),
    ("SA34", "C.5", "Auth API", "API: developer webhooks", "AUTO", "/api/developer/webhooks"),
    ("SA35", "C.6", "Auth API", "API: developer referrals", "AUTO", "/api/developer/referrals"),
    ("SA36", "C.7", "Auth API", "API: auth developer me", "AUTO", "/api/auth/developer/me"),
]


def create_dashboard(wb: Workbook):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "10B981"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "SettleGrid Smoke Test Dashboard"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="10B981")

    ws["A3"] = "Last Run:"
    ws["B3"] = "Not yet run"
    ws["A3"].font = Font(bold=True)

    # Summary counters
    ws["A5"] = "Overall Summary"
    ws["A5"].font = Font(bold=True, size=14, color="10B981")

    labels = ["Total", "Pass", "Fail", "Skip", "Pending"]
    for i, label in enumerate(labels):
        row = 6 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = len(TESTS) if label == "Total" else (len(TESTS) if label == "Pending" else 0)
        ws[f"B{row}"].font = Font(size=12)

    # Type breakdown
    ws["A13"] = "By Type"
    ws["A13"].font = Font(bold=True, size=14, color="10B981")

    type_counts = {"AUTO": 0, "SEMI": 0, "MANUAL": 0}
    for t in TESTS:
        type_counts[t[4]] += 1

    for i, (typ, count) in enumerate(type_counts.items()):
        row = 14 + i
        ws[f"A{row}"] = typ
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = TYPE_FILLS[typ]
        ws[f"B{row}"] = count
        ws[f"C{row}"] = "0 passed"

    # Phase breakdown
    ws["A19"] = "By Phase"
    ws["A19"].font = Font(bold=True, size=14, color="10B981")

    phases = {}
    for t in TESTS:
        phase = t[1].split(".")[0]
        cat = t[2]
        key = f"Phase {phase}: {cat}"
        if key not in phases:
            phases[key] = 0
        phases[key] += 1

    ws["A20"] = "Phase"
    ws["B20"] = "Tests"
    ws["C20"] = "Pass"
    ws["D20"] = "Fail"
    for cell in [ws["A20"], ws["B20"], ws["C20"], ws["D20"]]:
        cell.font = Font(bold=True)

    for i, (phase, count) in enumerate(phases.items()):
        row = 21 + i
        ws[f"A{row}"] = phase
        ws[f"B{row}"] = count
        ws[f"C{row}"] = 0
        ws[f"D{row}"] = 0

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12


def create_tests_sheet(wb: Workbook):
    ws = wb.create_sheet("Tests")
    ws.sheet_properties.tabColor = "10B981"

    # Headers
    headers = [
        ("ID", 6), ("Phase", 8), ("Category", 22), ("Test Name", 50),
        ("Type", 8), ("Status", 10), ("Endpoint", 35), ("HTTP Status", 12),
        ("Duration (ms)", 14), ("Error", 40), ("Last Run", 20), ("Notes", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, test in enumerate(TESTS, 2):
        test_id, phase, category, name, typ, endpoint = test
        values = [test_id, phase, category, name, typ, "PENDING", endpoint, "", "", "", "", ""]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

            # Type-based row fill
            if col_idx not in (6,):
                cell.fill = TYPE_FILLS.get(typ, PatternFill(fill_type=None))

            # Status cell fill
            if col_idx == 6:
                cell.fill = STATUS_FILLS.get(value, PatternFill(fill_type=None))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(TESTS) + 1}"


def main():
    wb = Workbook()
    create_dashboard(wb)
    create_tests_sheet(wb)
    wb.save(OUTPUT)
    print(f"✓ Created {OUTPUT}")
    print(f"  {len(TESTS)} tests across {len(set(t[1].split('.')[0] for t in TESTS))} phases")
    auto = sum(1 for t in TESTS if t[4] == "AUTO")
    semi = sum(1 for t in TESTS if t[4] == "SEMI")
    manual = sum(1 for t in TESTS if t[4] == "MANUAL")
    print(f"  AUTO: {auto} | SEMI: {semi} | MANUAL: {manual}")


if __name__ == "__main__":
    main()
