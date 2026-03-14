#!/usr/bin/env python3
"""Sync smoke-results.json → SMOKE_TEST_TRACKER.xlsx for SettleGrid."""

import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
RESULTS_FILE = os.path.join(PROJECT_DIR, "smoke-results.json")
TRACKER_FILE = os.path.join(PROJECT_DIR, "SMOKE_TEST_TRACKER.xlsx")

STATUS_FILLS = {
    "PASS": PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid"),
    "FAIL": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
    "SKIP": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
    "PENDING": PatternFill(fill_type=None),
}


def main():
    if not os.path.exists(RESULTS_FILE):
        print(f"No results file found: {RESULTS_FILE}")
        sys.exit(1)

    if not os.path.exists(TRACKER_FILE):
        print(f"No tracker file found: {TRACKER_FILE}")
        print("Run: python3 scripts/create-smoke-tracker.py first")
        sys.exit(1)

    with open(RESULTS_FILE) as f:
        data = json.load(f)

    results = {r["id"]: r for r in data.get("results", [])}
    timestamp = data.get("timestamp", "")

    wb = load_workbook(TRACKER_FILE)
    ws_tests = wb["Tests"]
    ws_dash = wb["Dashboard"]

    updated = 0
    pass_count = 0
    fail_count = 0
    skip_count = 0
    pending_count = 0

    # Update Tests sheet
    for row in range(2, ws_tests.max_row + 1):
        test_id = ws_tests.cell(row=row, column=1).value
        if not test_id:
            continue

        if test_id in results:
            r = results[test_id]
            status = r.get("status", "PENDING")
            ws_tests.cell(row=row, column=6, value=status)
            ws_tests.cell(row=row, column=6).fill = STATUS_FILLS.get(status, PatternFill(fill_type=None))
            ws_tests.cell(row=row, column=6).font = Font(bold=True)
            ws_tests.cell(row=row, column=8, value=r.get("http_status", ""))
            ws_tests.cell(row=row, column=9, value=r.get("duration_ms", 0))
            ws_tests.cell(row=row, column=10, value=r.get("error", ""))
            ws_tests.cell(row=row, column=11, value=r.get("timestamp", ""))
            updated += 1

        current_status = ws_tests.cell(row=row, column=6).value or "PENDING"
        if current_status == "PASS":
            pass_count += 1
        elif current_status == "FAIL":
            fail_count += 1
        elif current_status == "SKIP":
            skip_count += 1
        else:
            pending_count += 1

    # Update Dashboard
    total = pass_count + fail_count + skip_count + pending_count
    ws_dash["B3"] = timestamp
    ws_dash["B6"] = total
    ws_dash["B7"] = pass_count
    ws_dash["B8"] = fail_count
    ws_dash["B9"] = skip_count
    ws_dash["B10"] = pending_count

    # Update type breakdown
    type_pass = {"AUTO": 0, "SEMI": 0, "MANUAL": 0}
    for row in range(2, ws_tests.max_row + 1):
        typ = ws_tests.cell(row=row, column=5).value
        status = ws_tests.cell(row=row, column=6).value
        if typ and status == "PASS":
            type_pass[typ] = type_pass.get(typ, 0) + 1

    for i, typ in enumerate(["AUTO", "SEMI", "MANUAL"]):
        ws_dash.cell(row=14 + i, column=3, value=f"{type_pass[typ]} passed")

    # Update phase breakdown
    phase_stats = {}
    for row in range(2, ws_tests.max_row + 1):
        phase_cell = ws_tests.cell(row=row, column=2).value
        status = ws_tests.cell(row=row, column=6).value
        category = ws_tests.cell(row=row, column=3).value
        if phase_cell:
            phase_num = str(phase_cell).split(".")[0]
            key = f"Phase {phase_num}: {category}"
            if key not in phase_stats:
                phase_stats[key] = {"pass": 0, "fail": 0}
            if status == "PASS":
                phase_stats[key]["pass"] += 1
            elif status == "FAIL":
                phase_stats[key]["fail"] += 1

    for row in range(21, ws_dash.max_row + 1):
        phase_name = ws_dash.cell(row=row, column=1).value
        if phase_name and phase_name in phase_stats:
            ws_dash.cell(row=row, column=3, value=phase_stats[phase_name]["pass"])
            ws_dash.cell(row=row, column=4, value=phase_stats[phase_name]["fail"])

    wb.save(TRACKER_FILE)
    print(f"✓ Updated {updated} tests: {pass_count} passed, {fail_count} failed, {skip_count} skipped, {pending_count} pending")


if __name__ == "__main__":
    main()
